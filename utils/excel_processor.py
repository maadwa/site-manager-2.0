import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Optional, Dict, Any
import os

class ExcelProcessor:
    """Handles Excel file processing and manipulation"""
    
    def __init__(self):
        pass
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """Get list of sheet names in an Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        
        except Exception as e:
            print(f"Error getting sheet names: {str(e)}")
            return []
    
    def load_sheet_data(self, file_path: str, sheet_name: str) -> Optional[pd.DataFrame]:
        """Load data from a specific sheet"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Clean the data
            df = self._clean_dataframe(df)
            
            return df
        
        except Exception as e:
            print(f"Error loading sheet data: {str(e)}")
            return None
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and prepare dataframe for analysis"""
        try:
            # Remove completely empty rows and columns
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            # Convert date columns
            for col in df.columns:
                if 'date' in col.lower() or 'time' in col.lower():
                    try:
                        df[col] = pd.to_datetime(df[col], errors='ignore')
                    except:
                        pass
            
            # Convert numeric columns
            for col in df.columns:
                if df[col].dtype == 'object':
                    # Try to convert to numeric
                    numeric_series = pd.to_numeric(df[col], errors='ignore')
                    if not numeric_series.equals(df[col]):
                        df[col] = numeric_series
            
            return df
        
        except Exception as e:
            print(f"Error cleaning dataframe: {str(e)}")
            return df
    
    def create_statistics_sheet(self, file_path: str, data: pd.DataFrame, selected_columns: List[str]) -> bool:
        """Create or update Statistics sheet in the Excel file"""
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)
            
            # Check if Statistics sheet exists, if not create it
            if 'Statistics' in workbook.sheetnames:
                # Remove existing Statistics sheet
                del workbook['Statistics']
            
            # Create new Statistics sheet
            stats_sheet = workbook.create_sheet('Statistics')
            
            # Add header
            stats_sheet['A1'] = 'Construction Project Statistics'
            stats_sheet['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            
            # Add selected columns information
            row = 3
            stats_sheet[f'A{row}'] = 'Selected Columns for Analysis:'
            stats_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            
            row += 1
            for col in selected_columns:
                stats_sheet[f'A{row}'] = f"• {col}"
                row += 1
            
            # Add basic statistics
            row += 2
            stats_sheet[f'A{row}'] = 'Basic Statistics:'
            stats_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
            
            row += 1
            stats_sheet[f'A{row}'] = 'Total Records:'
            stats_sheet[f'B{row}'] = len(data)
            
            row += 1
            stats_sheet[f'A{row}'] = 'Total Columns:'
            stats_sheet[f'B{row}'] = len(data.columns)
            
            # Add column-wise statistics for numeric columns
            numeric_cols = data.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                row += 2
                stats_sheet[f'A{row}'] = 'Numeric Column Statistics:'
                stats_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
                
                row += 1
                # Headers
                stats_sheet[f'A{row}'] = 'Column'
                stats_sheet[f'B{row}'] = 'Count'
                stats_sheet[f'C{row}'] = 'Mean'
                stats_sheet[f'D{row}'] = 'Min'
                stats_sheet[f'E{row}'] = 'Max'
                stats_sheet[f'F{row}'] = 'Sum'
                
                # Make headers bold
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    stats_sheet[f'{col}{row}'].font = openpyxl.styles.Font(bold=True)
                
                # Add statistics for each numeric column
                for col in numeric_cols:
                    if col in selected_columns or not selected_columns:
                        row += 1
                        stats_sheet[f'A{row}'] = col
                        stats_sheet[f'B{row}'] = data[col].count()
                        stats_sheet[f'C{row}'] = round(data[col].mean(), 2) if data[col].count() > 0 else 0
                        stats_sheet[f'D{row}'] = data[col].min() if data[col].count() > 0 else 0
                        stats_sheet[f'E{row}'] = data[col].max() if data[col].count() > 0 else 0
                        stats_sheet[f'F{row}'] = data[col].sum() if data[col].count() > 0 else 0
            
            # Add note about graph generation
            row += 3
            stats_sheet[f'A{row}'] = 'Note: Graphs are generated in the web application based on selected columns.'
            stats_sheet[f'A{row}'].font = openpyxl.styles.Font(italic=True)
            
            row += 1
            stats_sheet[f'A{row}'] = 'Use the PowerPoint export feature to generate presentation slides.'
            stats_sheet[f'A{row}'].font = openpyxl.styles.Font(italic=True)
            
            # Save the workbook
            workbook.save(file_path)
            workbook.close()
            
            return True
        
        except Exception as e:
            print(f"Error creating statistics sheet: {str(e)}")
            return False
    
    def get_column_info(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """Get information about columns in a sheet"""
        try:
            df = self.load_sheet_data(file_path, sheet_name)
            
            if df is None:
                return {}
            
            column_info = {}
            
            for col in df.columns:
                column_info[col] = {
                    'dtype': str(df[col].dtype),
                    'non_null_count': df[col].count(),
                    'null_count': df[col].isnull().sum(),
                    'unique_count': df[col].nunique(),
                    'is_numeric': pd.api.types.is_numeric_dtype(df[col])
                }
                
                if pd.api.types.is_numeric_dtype(df[col]):
                    column_info[col].update({
                        'mean': df[col].mean(),
                        'min': df[col].min(),
                        'max': df[col].max(),
                        'std': df[col].std()
                    })
            
            return column_info
        
        except Exception as e:
            print(f"Error getting column info: {str(e)}")
            return {}
    
    def has_statistics_sheet(self, file_path: str) -> bool:
        """Check if file has a Statistics sheet"""
        try:
            sheet_names = self.get_sheet_names(file_path)
            return 'Statistics' in sheet_names
        except:
            return False
